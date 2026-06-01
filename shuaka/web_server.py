"""
本地 Web 服务器
- 平板叫号大屏页面
- 签到数据 JSON API
- 后台管理面板 (/admin) — 需登录
- 设置管理 + Logo 上传 — 需管理员权限
- 支持内网/外网双访问
- 支持百度网盘多机数据同步
"""

import os
import sys
import json
import time
import hmac
import hashlib
import socket
import secrets
import threading
import uuid
import glob as _glob
from datetime import datetime
from functools import wraps
from flask import Flask, jsonify, request, send_from_directory, redirect

from auth_manager import is_activated, get_machine_code, verify_license, save_license, get_license_info
from local_mirror import mirror_critical_data

# 路径常量
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TABLET_DIR = os.path.join(BASE_DIR, "tablet")

# 版本号
def _load_version():
    vf = os.path.join(BASE_DIR, "VERSION")
    if os.path.exists(vf):
        with open(vf, "r", encoding="utf-8") as f:
            return f.read().strip()
    return "v1.0.0.1"
VERSION = _load_version()

# 数据同步目录：固定使用项目下的 BaiduSyncdisk/ 目录
def _get_sync_root():
    from platform_utils import get_sync_data_dir
    return get_sync_data_dir(base_dir=BASE_DIR)

SYNC_ROOT = _get_sync_root()

UPLOADS_DIR = os.path.join(SYNC_ROOT, "uploads")
SETTINGS_FILE = os.path.join(SYNC_ROOT, "settings.json")
USERS_FILE = os.path.join(SYNC_ROOT, "users.json")
# 本机专属配置路径（存本地，不跟 SMB/云同步走）
def _get_machine_file():
    if os.name == 'nt':
        local_dir = os.path.join(os.environ.get('LOCALAPPDATA', os.path.expanduser('~')), 'shuaka')
    else:
        local_dir = os.path.join(os.path.expanduser('~'), '.shuaka')
    os.makedirs(local_dir, exist_ok=True)
    return os.path.join(local_dir, 'machine.json')

MACHINE_FILE = _get_machine_file()

os.makedirs(UPLOADS_DIR, exist_ok=True)
BACKUP_DIR = os.path.join(SYNC_ROOT, "备份")
RECYCLE_DIR = os.path.join(SYNC_ROOT, "回收站")
EVENTS_DIR = os.path.join(SYNC_ROOT, "events")
os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs(RECYCLE_DIR, exist_ok=True)
os.makedirs(EVENTS_DIR, exist_ok=True)

# 本机唯一标识（存 machine.json，不参与云同步），延迟初始化
_MID = None


def _get_machine_id():
    global _MID
    if _MID:
        return _MID
    ms = load_machine_settings()
    mid = ms.get("machine_id", "")
    if not mid:
        mid = uuid.uuid4().hex[:8]
        save_machine_settings({"machine_id": mid})
    _MID = mid
    return mid

# Flask app
app = Flask(__name__, static_folder=TABLET_DIR, static_url_path="")
app.config["SECRET_KEY"] = secrets.token_hex(32)

# 禁止浏览器缓存，确保每次都是最新版本
@app.after_request
def _no_cache(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

# ---------- 授权激活 ----------

_ACTIVATION_WHITELIST = {"/activate", "/api/activate", "/api/license-status"}


@app.before_request
def _check_activation():
    if request.path in _ACTIVATION_WHITELIST:
        return None
    if request.path.startswith("/static") or request.path.startswith("/uploads"):
        return None
    if not is_activated():
        if request.path.startswith("/api/"):
            return jsonify({"ok": False, "error": "软件未激活", "activate_url": "/activate"}), 403
        return redirect("/activate")


@app.route("/activate")
def activate_page():
    return send_from_directory(TABLET_DIR, "activate.html")


@app.route("/api/license-status")
def api_license_status():
    return jsonify(get_license_info())


@app.route("/api/activate", methods=["POST"])
def api_activate():
    data = request.get_json() or {}
    raw = data.get("code", "").strip()
    if not raw:
        return jsonify({"ok": False, "error": "请输入授权码"}), 400
    # 支持组合格式: XXXX-XXXX-XXXX-XXXX-YYYYMMDD 或分开传 code + expiry
    parts = raw.replace("-", " ").split()
    if len(parts) >= 5 and len(parts[4]) == 8:
        code = "-".join(parts[:4])
        expiry = parts[4]
    else:
        code = raw
        expiry = data.get("expiry", "").strip()
    if not code:
        return jsonify({"ok": False, "error": "授权码格式不正确"}), 400
    if not expiry or len(expiry) != 8:
        return jsonify({"ok": False, "error": "授权码格式不正确（缺少有效期）"}), 400
    if verify_license(code, expiry):
        save_license(code, expiry)
        return jsonify({"ok": True, "message": "激活成功"})
    return jsonify({"ok": False, "error": "授权码无效"}), 403


@app.route("/api/gen-license", methods=["POST"])
def api_gen_license():
    """生成授权码（需验证密码）"""
    data = request.get_json() or {}
    machine_code = data.get("machine_code", "").strip().upper()
    months = int(data.get("months", 1))
    pwd = data.get("password", "").strip()

    if pwd != "liaowei88":
        return jsonify({"ok": False, "error": "密码错误"}), 403
    if not machine_code or len(machine_code.replace("-", "")) < 8:
        return jsonify({"ok": False, "error": "请输入有效的机器码"}), 400
    if months not in (1, 3, 6, 12):
        return jsonify({"ok": False, "error": "有效期仅支持 1/3/6/12 个月"}), 400

    try:
        from auth_manager import generate_license
        from datetime import timedelta as _td
        lic, exp = generate_license(machine_code.replace("-", ""), months)
        code = f"{lic}-{exp}"
        expiry = (datetime.now() + _td(days=months * 30)).strftime("%Y-%m-%d")
        return jsonify({
            "ok": True,
            "license": code,
            "expiry": expiry,
            "months": months,
            "machine_code": machine_code
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# 全局引用
_excel_manager = None
_ngrok_url = None
_voice = None
_timer_manager = None


def set_voice_broadcaster(voice):
    global _voice
    _voice = voice


def set_timer_manager(tm):
    global _timer_manager
    _timer_manager = tm


def _apply_voice_settings(settings):
    """将语音设置实时同步到运行中的 VoiceBroadcaster"""
    global _voice
    if _voice and "voice" in settings:
        _voice.update_config(settings["voice"])


def _apply_timer_settings(settings):
    """将计时设置实时同步到运行中的 TimerManager"""
    global _timer_manager
    if _timer_manager and "timer" in settings:
        t = settings["timer"]
        if "remind_minutes" in t:
            _timer_manager.update_remind_minutes(t["remind_minutes"])


def _speak(text):
    if _voice and text:
        _voice.speak(text)

# 监控状态（跨模块共享）
_monitor_state = {
    "card_reader": {
        "enabled": False,       # 本机是否启用了读卡器
        "online": False,        # 读卡器是否实际在工作
        "last_read": None,      # ISO timestamp
        "last_name": "",
        "last_id": "",
        "total_reads": 0
    },
    "started_at": None,         # 系统启动时间
    "last_sync": "",            # 上次同步时间
    "sync_result": ""           # 同步结果描述
}


def set_card_reader_enabled(enabled=True):
    _monitor_state["card_reader"]["enabled"] = enabled
    if not enabled:
        _monitor_state["card_reader"]["online"] = False


def set_monitor_card_online(online=True):
    _monitor_state["card_reader"]["online"] = online


def set_monitor_card_read(name, id_number):
    _monitor_state["card_reader"]["last_read"] = datetime.now().isoformat()
    _monitor_state["card_reader"]["last_name"] = name
    _monitor_state["card_reader"]["last_id"] = id_number[:4] + "****" + id_number[-4:] if len(id_number) == 18 else id_number
    _monitor_state["card_reader"]["total_reads"] += 1


def set_monitor_started():
    _monitor_state["started_at"] = datetime.now().isoformat()

# 登录令牌存储 {token: {username, role, expires}}
_sessions = {}
SESSIONS_FILE = os.path.join(SYNC_ROOT, "sessions.json")

def _load_sessions():
    """从文件恢复登录会话"""
    global _sessions
    try:
        if os.path.exists(SESSIONS_FILE):
            with open(SESSIONS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            now = time.time()
            # 只恢复未过期的会话
            _sessions = {k: v for k, v in data.items() if v.get("expires", 0) > now}
            if _sessions:
                print(f"[session] 已恢复 {len(_sessions)} 个有效会话", flush=True)
    except Exception as e:
        print(f"[session] 加载失败: {e}", flush=True)

def _save_sessions():
    """持久化登录会话到文件"""
    try:
        # 清理过期会话
        now = time.time()
        active = {k: v for k, v in _sessions.items() if v.get("expires", 0) > now}
        _sessions.clear(); _sessions.update(active)
        with open(SESSIONS_FILE, "w", encoding="utf-8") as f:
            json.dump(_sessions, f, ensure_ascii=False)
    except Exception as e:
        print(f"[session] 保存失败: {e}", flush=True)


def set_excel_manager(mgr):
    global _excel_manager
    _excel_manager = mgr


def set_ngrok_url(url):
    global _ngrok_url
    _ngrok_url = url


def detect_ngrok_url():
    """从本地 ngrok API 自动获取外网地址"""
    global _ngrok_url
    try:
        import urllib.request, json as _json
        req = urllib.request.Request('http://127.0.0.1:4040/api/tunnels')
        with urllib.request.urlopen(req, timeout=3) as resp:
            data = _json.loads(resp.read())
        for t in data.get('tunnels', []):
            if t.get('proto') == 'https':
                _ngrok_url = t['public_url']
                print(f"[ngrok] 自动检测到外网地址: {_ngrok_url}", flush=True)
                return _ngrok_url
    except Exception:
        pass
    return None


# ---------- 用户管理 ----------

def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"admin": {"password": "admin123", "role": "admin", "name": "管理员"}}


def save_users(data):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    mirror_critical_data("users")


def check_auth(request):
    """验证请求的登录令牌，返回用户信息或 None"""
    token = request.headers.get("X-Auth-Token") or request.cookies.get("signin_token")
    if not token:
        return None
    session = _sessions.get(token)
    if not session:
        return None
    if time.time() > session["expires"]:
        del _sessions[token]
        return None
    return session


def require_auth(f):
    """装饰器：需要登录"""
    @wraps(f)
    def wrapper(*a, **kw):
        user = check_auth(request)
        if not user:
            return jsonify({"ok": False, "error": "请先登录"}), 401
        request.current_user = user
        return f(*a, **kw)
    return wrapper


def require_admin(f):
    """装饰器：需要管理员权限"""
    @wraps(f)
    def wrapper(*a, **kw):
        user = check_auth(request)
        if not user:
            return jsonify({"ok": False, "error": "请先登录"}), 401
        if user.get("role") != "admin":
            return jsonify({"ok": False, "error": "需要管理员权限"}), 403
        request.current_user = user
        return f(*a, **kw)
    return wrapper


# ---------- 登录 API ----------

@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "error": "无效请求"}), 400

    username = data.get("username", "").strip()
    password = data.get("password", "")

    users = load_users()
    user = users.get(username)
    if not user or user.get("password") != password:
        return jsonify({"ok": False, "error": "用户名或密码错误"}), 401

    # 生成令牌
    token = secrets.token_hex(32)
    _sessions[token] = {
        "username": username,
        "role": user.get("role", "user"),
        "name": user.get("name", username),
        "expires": time.time() + 86400 * 7  # 7天有效
    }
    _save_sessions()

    return jsonify({
        "ok": True,
        "token": token,
        "user": {"username": username, "role": user.get("role"), "name": user.get("name")}
    })


@app.route("/api/logout", methods=["POST"])
def api_logout():
    token = request.headers.get("X-Auth-Token") or request.cookies.get("signin_token")
    if token and token in _sessions:
        del _sessions[token]
    return jsonify({"ok": True})


@app.route("/api/me")
def api_me():
    user = check_auth(request)
    if not user:
        return jsonify({"ok": False}), 401
    return jsonify({"ok": True, "user": {
        "username": user["username"],
        "role": user["role"],
        "name": user["name"]
    }})


# ---------- 用户管理 API (管理员) ----------

@app.route("/api/users", methods=["GET"])
@require_admin
def api_list_users():
    users = load_users()
    result = {}
    for u, info in users.items():
        result[u] = {"role": info.get("role"), "name": info.get("name")}
    return jsonify({"ok": True, "users": result})


@app.route("/api/users", methods=["POST"])
@require_admin
def api_add_user():
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "error": "无效数据"}), 400

    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    role = data.get("role", "user")

    if not username or not password:
        return jsonify({"ok": False, "error": "用户名和密码不能为空"}), 400

    users = load_users()
    if username in users:
        return jsonify({"ok": False, "error": "用户已存在"}), 400

    users[username] = {"password": password, "role": role, "name": data.get("name", username)}
    save_users(users)
    return jsonify({"ok": True})


@app.route("/api/users/<username>", methods=["DELETE"])
@require_admin
def api_delete_user(username):
    if username == "admin":
        return jsonify({"ok": False, "error": "不能删除admin账号"}), 400
    users = load_users()
    if username not in users:
        return jsonify({"ok": False, "error": "用户不存在"}), 404
    del users[username]
    save_users(users)
    return jsonify({"ok": True})


# ---------- 设置管理 ----------

def load_settings():
    base = {}
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            base = json.load(f)
    # 合并本机专属配置（不参与云同步）
    if os.path.exists(MACHINE_FILE):
        with open(MACHINE_FILE, "r", encoding="utf-8") as f:
            machine = json.load(f)
            # 只覆盖 machine 里声明的顶级 key
            for k in machine:
                if k in base and isinstance(base[k], dict) and isinstance(machine[k], dict):
                    base[k] = deep_merge(base[k], machine[k])
                else:
                    base[k] = machine[k]
    return base

def load_machine_settings():
    """加载本机专属配置"""
    if os.path.exists(MACHINE_FILE):
        with open(MACHINE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_machine_settings(data):
    """保存本机专属配置（merge 模式）"""
    existing = load_machine_settings()
    merged = deep_merge(existing, data)
    with open(MACHINE_FILE, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
    # 同步到内存状态
    if "card_reader" in data:
        cr = data["card_reader"]
        if "enabled" in cr:
            set_card_reader_enabled(cr["enabled"])


def deep_merge(base, override):
    result = base.copy()
    for k, v in override.items():
        if k in result and isinstance(result[k], dict) and isinstance(v, dict):
            result[k] = deep_merge(result[k], v)
        else:
            result[k] = v
    return result


def save_settings(data, merge=True):
    if merge:
        existing = load_settings()
        data = deep_merge(existing, data)
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    mirror_critical_data("settings")


# ---------- 多机实时联动（通过云同步文件夹传递事件）----------

_processed_events = set()
_event_poll_thread = None


def _write_event(action, name=None, text=None, card=None):
    """写入事件文件到云同步目录，其他机器轮询到后处理"""
    payload = {
        "machine": _get_machine_id(),
        "action": action,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    if name:
        payload["name"] = name
    if text:
        payload["text"] = text
    if card:
        payload["card"] = card
    fname = f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{_get_machine_id()}.json"
    fpath = os.path.join(EVENTS_DIR, fname)
    try:
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)
        print(f"[联动] 写入事件: {action} → {fname}", flush=True)
    except Exception:
        pass


def _process_events():
    """扫描并处理其他机器写入的事件文件"""
    global _processed_events
    try:
        files = sorted(_glob.glob(os.path.join(EVENTS_DIR, "*.json")))
        now = time.time()
        for fpath in files:
            fname = os.path.basename(fpath)
            # 清理超过 1 小时的旧事件
            try:
                if now - os.path.getmtime(fpath) > 3600:
                    os.remove(fpath)
                    _processed_events.add(fname)
                    continue
            except Exception:
                pass
            if fname in _processed_events:
                continue
            try:
                with open(fpath, "r", encoding="utf-8") as f:
                    data = json.load(f)
            except Exception:
                _processed_events.add(fname)
                continue

            if data.get("machine") == _get_machine_id():
                _processed_events.add(fname)
                continue

            action = data.get("action", "")
            print(f"[联动] 收到来自 {data.get('machine','?')} 的事件: {action}", flush=True)

            if action in ("refresh", "signin"):
                if _excel_manager:
                    try:
                        _excel_manager._load_cache_from_files()
                    except Exception:
                        pass

            if action == "signin":
                name = data.get("name", "")
                if name and _timer_manager:
                    _timer_manager.add_timer(name, "")

            # 语音类事件不在此删除，留给前端轮询播放
            if action in ("speak", "signin"):
                _processed_events.add(fname)
                continue

            try:
                os.remove(fpath)
            except Exception:
                pass
            _processed_events.add(fname)

        if len(_processed_events) > 500:
            existing = set(os.path.basename(p) for p in _glob.glob(os.path.join(EVENTS_DIR, "*.json")))
            _processed_events = {n for n in _processed_events if n in existing}
    except Exception:
        pass


def _start_event_polling():
    """启动后台事件轮询线程（每3秒扫描一次云同步目录）"""
    global _event_poll_thread

    def _poll_loop():
        while True:
            time.sleep(3)
            _process_events()

    _event_poll_thread = threading.Thread(target=_poll_loop, daemon=True)
    _event_poll_thread.start()


# ---------- 页面路由 ----------

@app.route("/")
def index():
    return send_from_directory(TABLET_DIR, "index.html")


@app.route("/login")
def login_page():
    return send_from_directory(TABLET_DIR, "login.html")


@app.route("/admin")
def admin_page():
    # admin.html 自己会检查登录状态，未登录跳转 /login
    return send_from_directory(TABLET_DIR, "admin.html")


# ---------- 数据 API（无需登录） ----------

@app.route("/api/signins")
def api_signins():
    if _excel_manager is None:
        return jsonify([])
    return jsonify(_excel_manager.get_all_records())


@app.route("/api/status")
def api_status():
    if _excel_manager is None:
        return jsonify({"status": "未就绪", "record_count": 0})
    return jsonify({
        "status": "运行中",
        "version": VERSION,
        "record_count": len(_excel_manager.get_all_records()),
        "today_count": len(_excel_manager.get_today_records()),
        "ngrok_url": _ngrok_url
    })


# ---------- 设置 API ----------

@app.route("/api/settings", methods=["GET"])
def get_settings():
    return jsonify(load_settings())


@app.route("/api/settings", methods=["POST"])
@require_admin
def update_settings():
    data = request.get_json()
    if data:
        # 保护跑马灯：如果提交的 marquees 全是空的，保留原值不覆盖
        if "marquees" in data:
            mqs = data["marquees"]
            all_empty = all(
                not (m.get("enabled") and m.get("text", "").strip())
                for m in mqs
            )
            if all_empty:
                existing = load_settings()
                if existing.get("marquees"):
                    data["marquees"] = existing["marquees"]
        save_settings(data, merge=True)
        # 实时同步语音/计时设置到运行中的实例
        _apply_voice_settings(data)
        _apply_timer_settings(data)
        # 同步地点名到 Excel 管理器
        if "display" in data and "location" in data["display"] and data["display"]["location"]:
            global _excel_manager
            if _excel_manager:
                _excel_manager.location = data["display"]["location"]
        _write_event("refresh")
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "无效数据"}), 400


# ---------- 本机设置 API（不参与云同步）----------

@app.route("/api/machine-settings", methods=["GET"])
def api_get_machine():
    settings = load_machine_settings()
    from platform_utils import get_baidu_sync_dir
    detected = get_baidu_sync_dir()
    settings["_detected_baidu"] = detected if os.path.isdir(detected) else ""
    return jsonify(settings)

@app.route("/api/machine-settings", methods=["POST"])
@require_admin
def api_save_machine():
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "error": "无效数据"}), 400
    # 处理路径检测请求（不保存）
    if "_check_path" in data:
        path = data["_check_path"]
        exists = os.path.isdir(path) if path else False
        return jsonify({"ok": True, "path_exists": exists})
    save_machine_settings(data)
    msg = "已保存"
    if "card_reader" in data:
        msg += " · 读卡器设置需重启服务生效"
    if "baidu_sync_dir" in data:
        msg += " · 同步目录需重启服务生效"
    return jsonify({"ok": True, "message": msg})


@app.route("/api/browse-folder", methods=["POST"])
@require_admin
def api_browse_folder():
    """弹出系统原生文件夹选择对话框，返回所选路径"""
    import subprocess
    try:
        if os.name == 'nt':
            ps_cmd = (
                'Add-Type -AssemblyName System.Windows.Forms; '
                '$f = New-Object System.Windows.Forms.FolderBrowserDialog; '
                '$f.Description = "选择百度云同步目录"; '
                'if ($f.ShowDialog() -eq "OK") { $f.SelectedPath } else { "" }'
            )
            result = subprocess.run(
                ['powershell', '-Command', ps_cmd],
                capture_output=True, text=True, timeout=30
            )
            path = result.stdout.strip()
        else:
            result = subprocess.run(
                ['osascript', '-e', 'POSIX path of (choose folder with prompt "选择百度云同步目录")'],
                capture_output=True, text=True, timeout=30
            )
            path = result.stdout.strip()

        if path and os.path.isdir(path):
            return jsonify({"ok": True, "path": path})
        return jsonify({"ok": False, "error": "未选择目录"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ---------- Logo 上传 ----------

@app.route("/api/upload/logo", methods=["POST"])
@require_admin
def upload_logo():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "未选择文件"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"ok": False, "error": "文件名为空"}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg"):
        return jsonify({"ok": False, "error": "仅支持 png/jpg/gif/webp/svg"}), 400

    filename = "logo" + ext
    filepath = os.path.join(UPLOADS_DIR, filename)

    for old_ext in (".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg"):
        old_path = os.path.join(UPLOADS_DIR, "logo" + old_ext)
        if old_path != filepath and os.path.exists(old_path):
            os.remove(old_path)

    file.save(filepath)

    return jsonify({"ok": True, "logo_url": f"/uploads/{filename}"})


@app.route("/api/location", methods=["POST"])
@require_admin
def update_location():
    """更新签到地点"""
    data = request.get_json()
    if not data or "location" not in data:
        return jsonify({"ok": False, "error": "缺少location参数"}), 400
    global _excel_manager
    if _excel_manager:
        _excel_manager.location = data["location"]
    # 同时保存到 settings
    settings = load_settings()
    settings.setdefault("display", {})["location"] = data["location"]
    save_settings(settings, merge=False)
    return jsonify({"ok": True, "location": data["location"]})


@app.route("/api/monitor")
def api_monitor():
    """返回系统监控状态：读卡器、Excel同步文件、百度云同步状态"""
    import glob as glob_mod
    from platform_utils import get_baidu_sync_dir, get_sync_data_dir

    sync_dir = get_sync_data_dir()

    result = {
        "version": VERSION,
        "card_reader": dict(_monitor_state["card_reader"]),
        "started_at": _monitor_state["started_at"],
        "sync": {
            "enabled": True,
            "provider": "BaiduSyncdisk/ 目录",
            "data_dir": sync_dir,
            "shared_items": ["签到记录", "备份", "回收站", "设置", "用户", "Logo"],
            "last_sync": _monitor_state.get("last_sync", ""),
            "sync_result": _monitor_state.get("sync_result", "")
        },
        "excel_dir": {
            "path": os.path.abspath(_excel_manager.excel_dir) if _excel_manager else "",
            "files": []
        }
    }
    if _excel_manager:
        for fp in sorted(glob_mod.glob(os.path.join(_excel_manager.excel_dir, "签到记录_*.xlsx"))):
            st = os.stat(fp)
            result["excel_dir"]["files"].append({
                "name": os.path.basename(fp),
                "size_kb": round(st.st_size / 1024, 1),
                "modified": datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            })
    return jsonify(result)


@app.route("/api/manual_signin", methods=["POST"])
def manual_signin():
    """网页手动签到"""
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "error": "无效数据"}), 400

    name = data.get("name", "").strip()
    id_number = data.get("id_number", "").strip() or "手动录入"
    if not name:
        return jsonify({"ok": False, "error": "请输入姓名"}), 400

    if _excel_manager is None:
        return jsonify({"ok": False, "error": "系统未就绪"}), 500

    from datetime import datetime
    record = _excel_manager.add_record(name, id_number)
    _speak(f"{name}，手动签到成功")
    _write_event("signin", name=name, text=f"{name}，手动签到成功")

    return jsonify({
        "ok": True,
        "record": record
    })


@app.route("/report")
def report_page():
    return send_from_directory(TABLET_DIR, "report.html")




# ---------- Excel 文件查看 ----------

@app.route("/api/excel/view/<path:filename>")
def api_excel_view(filename):
    """在浏览器中查看 Excel 签到文件内容"""
    if _excel_manager is None:
        return "<h3>系统未就绪</h3>", 500

    import os as _os
    safe_name = _os.path.basename(filename)
    filepath = _os.path.join(_os.path.abspath(_excel_manager.excel_dir), safe_name)

    if not _os.path.exists(filepath):
        return "<h3>文件不存在</h3>", 404

    try:
        from openpyxl import load_workbook as _lw
        wb = _lw(filepath)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return "<h3>文件为空</h3>"

        html = ['<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8">'
                '<meta name="viewport" content="width=device-width,initial-scale=1">'
                '<title>' + safe_name + '</title>'
                '<style>'
                '*{margin:0;padding:0;box-sizing:border-box}'
                'body{font-family:-apple-system,BlinkMacSystemFont,"PingFang SC","Microsoft YaHei",sans-serif;'
                'background:#0a0e14;color:#e2e6ec;padding:1rem}'
                'h2{color:#4f8fff;margin-bottom:1rem}'
                '.info{font-size:0.8rem;color:#8b95a5;margin-bottom:0.5rem}'
                'table{width:100%;border-collapse:collapse;background:#131820;border-radius:10px;overflow:hidden;font-size:0.85rem}'
                'th{background:#1a1f2b;color:#8b95a5;padding:0.6rem 0.4rem;text-align:center;font-weight:600;font-size:0.78rem;border-bottom:1px solid #1e2633}'
                'td{padding:0.5rem 0.4rem;text-align:center;border-bottom:1px solid rgba(255,255,255,0.04)}'
                'tr:nth-child(even){background:rgba(255,255,255,0.015)}'
                'tr:hover{background:rgba(79,143,255,0.06)}'
                '@media(max-width:600px){table{font-size:0.7rem}th,td{padding:0.35rem 0.15rem}}'
                '</style></head><body>'
                '<h2>📋 ' + safe_name + '</h2>'
                '<p class="info">共 ' + str(len(rows)-1) + ' 条记录</p>'
                '<table>']

        # 表头
        html.append('<thead><tr>')
        for cell in rows[0]:
            html.append('<th>' + (str(cell) if cell else '') + '</th>')
        html.append('</tr></thead><tbody>')

        # 数据行
        for row in rows[1:]:
            html.append('<tr>')
            for cell in row:
                html.append('<td>' + (str(cell) if cell is not None else '') + '</td>')
            html.append('</tr>')

        html.append('</tbody></table></body></html>')
        return '\n'.join(html)

    except Exception as e:
        return "<h3>读取失败: " + str(e) + "</h3>", 500


# ---------- 数据统计 API ----------

@app.route("/api/stats")
def api_stats():
    """签到统计报表"""
    if _excel_manager is None:
        return jsonify({})
    period = request.args.get("period", "week")
    records = _excel_manager.get_all_records()

    from datetime import datetime, timedelta
    now = datetime.now()
    today = now.strftime("%Y-%m-%d")

    result = {
        "total": len(records),
        "today": len([r for r in records if r["sign_time"].startswith(today)]),
        "location": _excel_manager.location if _excel_manager else "",
    }

    # 本周/本月
    monday = now - timedelta(days=now.weekday())
    result["this_week"] = len([r for r in records if r["sign_time"] >= monday.strftime("%Y-%m-%d")])
    result["this_month"] = len([r for r in records if r["sign_time"][:7] == now.strftime("%Y-%m")])

    # 按周期聚合
    labels, values = [], []
    if period == "day":
        for i in range(6, -1, -1):
            d = (now - timedelta(days=i)).strftime("%m-%d")
            cnt = len([r for r in records if r["sign_time"].startswith((now - timedelta(days=i)).strftime("%Y-%m-%d"))])
            labels.append(d); values.append(cnt)
    elif period == "month":
        for i in range(5, -1, -1):
            ym = (now.replace(day=1) - timedelta(days=i*30)).strftime("%Y-%m")
            cnt = len([r for r in records if r["sign_time"][:7] == ym])
            labels.append(ym); values.append(cnt)
    else:  # week
        weekdays = ["周一","周二","周三","周四","周五","周六","周日"]
        for i in range(6, -1, -1):
            d = (now - timedelta(days=i))
            cnt = len([r for r in records if r["sign_time"].startswith(d.strftime("%Y-%m-%d"))])
            labels.append(weekdays[d.weekday()]); values.append(cnt)

    result["labels"] = labels
    result["values"] = values

    # 时段分布
    hour_dist = {}
    for r in records:
        try:
            h = r["sign_time"][11:13]
            hour_dist[h] = hour_dist.get(h, 0) + 1
        except: pass
    result["hourly"] = [{"hour": int(h), "count": c} for h, c in sorted(hour_dist.items())]

    # 状态分布
    status_dist = {"正常等待": 0, "即将超时": 0, "已超时": 0}
    warn_min = load_settings().get("timer", {}).get("warning_minutes", 35)
    over_min = load_settings().get("timer", {}).get("remind_minutes", 40)
    for r in records:
        try:
            t = datetime.strptime(r["sign_time"], "%Y-%m-%d %H:%M:%S")
            waited = (now - t).total_seconds() / 60
            if waited >= over_min: status_dist["已超时"] += 1
            elif waited >= warn_min: status_dist["即将超时"] += 1
            else: status_dist["正常等待"] += 1
        except: pass
    result["status_dist"] = {k: v for k, v in status_dist.items() if v > 0}

    # 地点分布
    loc_dist = {}
    for r in records:
        loc = r.get("location", "未知")
        loc_dist[loc] = loc_dist.get(loc, 0) + 1
    result["location_dist"] = loc_dist

    return jsonify(result)


# ---------- 清除签到记录 ----------

@app.route("/api/clear_records", methods=["POST"])
@require_admin
def clear_records():
    """清除签到记录（先备份到回收站）"""
    if _excel_manager is None:
        return jsonify({"ok": False, "error": "系统未就绪"}), 500

    data = request.get_json() or {}
    mode = data.get("mode", "today")

    import glob, shutil
    from datetime import datetime
    excel_dir = _excel_manager.excel_dir
    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")

    if mode == "all":
        # 备份全部文件到回收站
        for f in glob.glob(os.path.join(excel_dir, "签到记录_*.xlsx")):
            try:
                bak_name = f"{now_str}_全部_{os.path.basename(f)}"
                shutil.copy2(f, os.path.join(RECYCLE_DIR, bak_name))
                os.remove(f)
            except: pass
        _excel_manager._records_cache = []
        _speak("全部签到记录已清除，已备份到回收站")
        _write_event("refresh")
        return jsonify({"ok": True, "message": "已清除全部签到记录（已备份到回收站）"})

    # 只清今天：备份后重建
    today = datetime.now().strftime("%Y-%m-%d")
    all_recs = _excel_manager.get_all_records()
    removed_recs = [r for r in all_recs if r["sign_time"].startswith(today)]
    keep = [r for r in all_recs if not r["sign_time"].startswith(today)]
    removed = len(removed_recs)

    if removed == 0:
        return jsonify({"ok": True, "message": "今日无记录需要清除"})

    # 备份今日被删除的记录到回收站 JSON
    import json as _json
    bak_path = os.path.join(RECYCLE_DIR, f"{now_str}_今日删除_{removed}条.json")
    with open(bak_path, "w", encoding="utf-8") as bf:
        _json.dump(removed_recs, bf, ensure_ascii=False, indent=2)

    # 重建所有 Excel 文件（不含今日记录）
    for f in glob.glob(os.path.join(excel_dir, "签到记录_*.xlsx")):
        try: os.remove(f)
        except: pass

    groups = {}
    for r in keep:
        groups.setdefault(r.get("location", "未知"), []).append(r)

    orig_loc = _excel_manager.location
    try:
        for loc, recs in groups.items():
            _excel_manager.location = loc
            for r in recs:
                t = datetime.strptime(r["sign_time"], "%Y-%m-%d %H:%M:%S")
                _excel_manager.add_record(r["name"], r["id_number"], t, r.get("status","等待中"), {"_recalled": r.get("_recalled",0)}, _rebuild=True)
    finally:
        _excel_manager.location = orig_loc

    _excel_manager._records_cache = keep
    _speak(f"已清除今日{removed}条记录，已备份到回收站")
    _write_event("refresh")
    return jsonify({"ok": True, "message": f"已清除今日 {removed} 条记录（已备份到回收站）"})


# ---------- 删除/编辑记录 ----------

@app.route("/api/delete_records", methods=["POST"])
def delete_records():
    """删除指定签到记录（备份到回收站）"""
    if _excel_manager is None:
        return jsonify({"ok": False, "error": "系统未就绪"}), 500
    data = request.get_json() or {}
    targets = data.get("targets", [])
    if not targets:
        return jsonify({"ok": False, "error": "请指定要删除的记录"}), 400

    import glob, json as _json
    from datetime import datetime
    all_recs = _excel_manager.get_all_records()
    target_set = set()
    for t in targets:
        target_set.add((t.get("seq"), t.get("location", "")))

    # 找出被删除的记录并备份
    deleted_recs = [r for r in all_recs if (r.get("seq"), r.get("location", "")) in target_set]
    if deleted_recs:
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        names = ','.join([r.get('name','?') for r in deleted_recs[:5]])
        bak_path = os.path.join(RECYCLE_DIR, f"{now_str}_删除_{names}.json")
        with open(bak_path, "w", encoding="utf-8") as bf:
            _json.dump(deleted_recs, bf, ensure_ascii=False, indent=2)

    keep = [r for r in all_recs if (r.get("seq"), r.get("location", "")) not in target_set]
    removed = len(all_recs) - len(keep)

    # 重建 Excel
    excel_dir = _excel_manager.excel_dir
    for f in glob.glob(os.path.join(excel_dir, "签到记录_*.xlsx")):
        try: os.remove(f)
        except Exception: pass

    groups = {}
    for r in keep:
        groups.setdefault(r.get("location", "未知"), []).append(r)

    orig_loc = _excel_manager.location
    try:
        for loc, recs in groups.items():
            _excel_manager.location = loc
            for r in recs:
                t = datetime.strptime(r["sign_time"], "%Y-%m-%d %H:%M:%S")
                _excel_manager.add_record(r["name"], r["id_number"], t, r.get("status","等待中"), {"_recalled": r.get("_recalled",0)}, _rebuild=True)
    finally:
        _excel_manager.location = orig_loc

    _excel_manager._records_cache = keep
    _speak(f"已删除{removed}条记录，已备份到回收站")
    _write_event("refresh")
    return jsonify({"ok": True, "message": f"已删除 {removed} 条记录（已备份到回收站）", "removed": removed})


@app.route("/api/update_record", methods=["POST"])
def update_record():
    """编辑签到记录（状态/重叫次数等）"""
    if _excel_manager is None:
        return jsonify({"ok": False, "error": "系统未就绪"}), 500
    data = request.get_json() or {}
    seq = data.get("seq")
    location = data.get("location", "")
    new_name = data.get("name", "").strip()
    new_id = data.get("id_number", "").strip()
    new_status = data.get("_set_status", "").strip()
    recalled = data.get("_recalled")

    if not seq:
        return jsonify({"ok": False, "error": "缺少必要参数"}), 400

    # 更新缓存（_recalled 等不存 Excel 的字段）
    for r in _excel_manager._records_cache:
        if str(r.get("seq")) == str(seq) and r.get("location", "") == location:
            if new_name: r["name"] = new_name
            if new_id: r["id_number"] = new_id
            if recalled is not None: r["_recalled"] = recalled
            break

    # 更新 Excel（直接搜 Excel 单元格，不依赖缓存）
    if new_status:
        ok = _excel_manager.update_status(seq, location, new_status)
        if not ok:
            # 文件不存在或记录不匹配，fallback 重建
            import glob
            all_recs = _excel_manager._records_cache
            excel_dir = _excel_manager.excel_dir
            for f in glob.glob(os.path.join(excel_dir, "签到记录_*.xlsx")):
                try: os.remove(f)
                except Exception: pass
            groups = {}
            for r in all_recs:
                groups.setdefault(r.get("location", "未知"), []).append(r)
            orig_loc = _excel_manager.location
            from datetime import datetime as dt
            try:
                for loc, recs in groups.items():
                    _excel_manager.location = loc
                    for r in recs:
                        t = dt.strptime(r["sign_time"], "%Y-%m-%d %H:%M:%S")
                        _excel_manager.add_record(r["name"], r["id_number"], t, r.get("status","等待中"), {"_recalled": r.get("_recalled",0)}, _rebuild=True)
            finally:
                _excel_manager.location = orig_loc
            _excel_manager._load_cache_from_files()

    peer_voice = data.get("_peer_voice", "").strip()
    if peer_voice:
        # 只有叫号操作才附带卡片信息（完成/过号应清屏，不附带卡片）
        card_info = None
        if new_status == "已叫号":
            card_info = {"seq": seq, "location": location}
            for r in _excel_manager._records_cache:
                if str(r.get("seq")) == str(seq) and r.get("location", "") == location:
                    card_info["name"] = r.get("name", "")
                    break
        _write_event("speak", text=peer_voice, card=card_info)
        _write_event("refresh")

    return jsonify({"ok": True, "message": "记录已更新"})


@app.route("/api/restore_record", methods=["POST"])
def restore_record():
    """恢复记录到等待队列：重置状态 + 签到时间（重排等候钟）"""
    if _excel_manager is None:
        return jsonify({"ok": False, "error": "系统未就绪"}), 500
    data = request.get_json() or {}
    seq = data.get("seq")
    location = data.get("location", "")
    if not seq:
        return jsonify({"ok": False, "error": "缺少必要参数"}), 400

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 更新缓存
    for r in _excel_manager._records_cache:
        if str(r.get("seq")) == str(seq) and r.get("location", "") == location:
            r["status"] = "等待中"
            r["sign_time"] = now_str
            break

    # 精准更新 Excel 状态列 + 时间列
    try:
        _excel_manager.update_status(seq, location, "等待中")
        # 同时更新签到时间
        from openpyxl import load_workbook as _lw
        filepath = _excel_manager._filepath(location)
        if os.path.exists(filepath):
            wb = _lw(filepath)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                if str(row[0].value) == str(seq):
                    row[3].value = now_str  # 签到时间列
                    wb.save(filepath)
                    break
            wb.close()
    except Exception:
        pass

    peer_voice = data.get("_peer_voice", "").strip()
    if peer_voice:
        _write_event("speak", text=peer_voice)
    _write_event("refresh")
    return jsonify({"ok": True, "restored_at": now_str})


@app.route("/api/swap_records", methods=["POST"])
def swap_records():
    """交换两条记录的签到时间（拖拽排序）"""
    if _excel_manager is None:
        return jsonify({"ok": False, "error": "系统未就绪"}), 500
    data = request.get_json() or {}
    a = data.get("a", {}); b = data.get("b", {})
    if not a or not b:
        return jsonify({"ok": False, "error": "缺少参数"}), 400
    rec_a = rec_b = None
    for r in _excel_manager._records_cache:
        if str(r.get("seq")) == str(a.get("seq")) and r.get("location", "") == a.get("loc", ""):
            rec_a = r
        if str(r.get("seq")) == str(b.get("seq")) and r.get("location", "") == b.get("loc", ""):
            rec_b = r
    if rec_a and rec_b:
        rec_a["sign_time"], rec_b["sign_time"] = rec_b["sign_time"], rec_a["sign_time"]
    import glob
    excel_dir = _excel_manager.excel_dir
    all_recs = _excel_manager._records_cache
    for f in glob.glob(os.path.join(excel_dir, "签到记录_*.xlsx")):
        try: os.remove(f)
        except Exception: pass
    groups = {}
    for r in all_recs:
        groups.setdefault(r.get("location", "未知"), []).append(r)
    orig_loc = _excel_manager.location
    from datetime import datetime as _dt
    try:
        for loc, recs in groups.items():
            _excel_manager.location = loc
            for r in recs:
                t = _dt.strptime(r["sign_time"], "%Y-%m-%d %H:%M:%S")
                _excel_manager.add_record(r["name"], r["id_number"], t, r.get("status","等待中"), {"_recalled": r.get("_recalled",0)}, _rebuild=True)
    finally:
        _excel_manager.location = orig_loc
    _excel_manager._load_cache_from_files()
    _write_event("refresh")
    return jsonify({"ok": True})


@app.route("/uploads/<filename>")
def serve_upload(filename):
    return send_from_directory(UPLOADS_DIR, filename)


# ---------- 网络工具 ----------

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(0)
        s.connect(("10.254.254.254", 1))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def get_all_ips():
    ips = set()
    try:
        for info in socket.getaddrinfo(socket.gethostname(), None):
            ip = info[4][0]
            if not ip.startswith("127.") and ":" not in ip:
                ips.add(ip)
    except Exception:
        pass
    ips.add(get_local_ip())
    return sorted(ips)



# ---------- 多机联动 API ----------

@app.route("/api/event", methods=["POST"])
def api_event():
    """HTTP 方式接收对端事件（云同步文件夹轮询为主，此接口作为补充）"""
    data = request.get_json()
    if not data:
        return jsonify({"ok": True})
    action = data.get("action", "")

    if action in ("speak", "signin"):
        text = data.get("text", "")
        if text:
            _speak(text)

    if action in ("refresh", "signin"):
        if _excel_manager:
            try:
                _excel_manager._load_cache_from_files()
            except Exception:
                pass

    if action == "signin":
        name = data.get("name", "")
        if name and _timer_manager:
            _timer_manager.add_timer(name, "")

    return jsonify({"ok": True})


@app.route("/api/sync-status")
def api_sync_status():
    """返回多机同步状态"""
    event_files = _glob.glob(os.path.join(EVENTS_DIR, "*.json"))
    pending = 0
    for fp in event_files:
        fname = os.path.basename(fp)
        if fname not in _processed_events:
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    d = json.load(f)
                if d.get("machine") != _get_machine_id():
                    pending += 1
            except Exception:
                pass
    return jsonify({
        "machine_id": _get_machine_id(),
        "events_dir": EVENTS_DIR,
        "pending_events": pending,
        "total_events": len(event_files),
    })


@app.route("/api/test-voice")
def api_test_voice():
    """测试语音播报"""
    if not _voice:
        return jsonify({"ok": False, "error": "语音播报未初始化"})
    _speak("多机联动语音测试")
    return jsonify({"ok": True, "message": "语音测试已发送"})


@app.route("/api/pending-voice")
def api_pending_voice():
    """轮询来自其他机器的待播语音，附带叫号卡片信息，返回后自动删除"""
    try:
        files = sorted(_glob.glob(os.path.join(EVENTS_DIR, "*.json")))
        for fpath in files:
            try:
                with open(fpath, "r", encoding="utf-8") as f:
                    d = json.load(f)
            except Exception:
                try:
                    os.remove(fpath)
                except Exception:
                    pass
                continue
            if d.get("machine") == _get_machine_id():
                continue
            if d.get("action") not in ("speak", "signin"):
                continue
            text = d.get("text", "")
            card = d.get("card") if d.get("action") == "speak" else None
            try:
                os.remove(fpath)
            except Exception:
                pass
            if text:
                return jsonify({"ok": True, "text": text, "card": card})
    except Exception:
        pass
    return jsonify({"ok": True, "text": "", "card": None})


# ---------- 备份管理 ----------

@app.route("/api/backup", methods=["GET"])
def api_backup_list():
    """列出所有备份文件"""
    import glob
    files = []
    for f in sorted(glob.glob(os.path.join(BACKUP_DIR, "*")), reverse=True):
        st = os.stat(f)
        files.append({
            "name": os.path.basename(f),
            "size_kb": round(st.st_size / 1024, 1),
            "modified": datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        })
    return jsonify({"ok": True, "files": files})


@app.route("/api/backup", methods=["POST"])
@require_admin
def api_backup_create():
    """手动备份：把当前所有签到记录打包备份"""
    import shutil, glob
    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_dir = _excel_manager.excel_dir if _excel_manager else "./签到记录"
    count = 0
    for f in glob.glob(os.path.join(excel_dir, "签到记录_*.xlsx")):
        try:
            bak_name = f"{now_str}_手动备份_{os.path.basename(f)}"
            shutil.copy2(f, os.path.join(BACKUP_DIR, bak_name))
            count += 1
        except: pass
    _speak(f"手动备份完成，共{count}个文件")
    return jsonify({"ok": True, "message": f"已备份 {count} 个文件"})


# ---------- 回收站管理 ----------

@app.route("/api/recycle", methods=["GET"])
def api_recycle_list():
    """列出回收站文件"""
    import glob
    files = []
    for f in sorted(glob.glob(os.path.join(RECYCLE_DIR, "*")), reverse=True):
        st = os.stat(f)
        files.append({
            "name": os.path.basename(f),
            "size_kb": round(st.st_size / 1024, 1),
            "modified": datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        })
    return jsonify({"ok": True, "files": files})


@app.route("/api/recycle/restore", methods=["POST"])
@require_admin
def api_recycle_restore():
    """从回收站恢复文件"""
    data = request.get_json() or {}
    filename = data.get("filename", "")
    if not filename:
        return jsonify({"ok": False, "error": "请指定文件名"}), 400
    src = os.path.join(RECYCLE_DIR, os.path.basename(filename))
    if not os.path.exists(src):
        return jsonify({"ok": False, "error": "文件不存在"}), 404
    import shutil
    excel_dir = _excel_manager.excel_dir if _excel_manager else "./签到记录"
    os.makedirs(excel_dir, exist_ok=True)
    # 如果是 JSON 文件（记录备份），还原到 Excel
    if src.endswith('.json'):
        import json as _json
        with open(src, "r", encoding="utf-8") as bf:
            recs = _json.load(bf)
        # 重建缓存
        if _excel_manager:
            existing = _excel_manager.get_all_records()
            for r in recs:
                _excel_manager.add_record(r["name"], r["id_number"],
                    datetime.strptime(r["sign_time"], "%Y-%m-%d %H:%M:%S"),
                    r.get("status", "等待中"),
                    {"_recalled": r.get("_recalled", 0)})
        try: os.remove(src)
        except: pass
    else:
        # Excel 备份文件直接恢复
        dst = os.path.join(excel_dir, filename.split("_手动备份_")[-1] if "_手动备份_" in filename else filename.split("_全部_")[-1] if "_全部_" in filename else filename)
        # 提取原始文件名
        parts = filename.split("_", 2)
        orig_name = parts[-1] if len(parts) > 2 else filename
        for prefix in ["手动备份_", "全部_"]:
            if orig_name.startswith(prefix):
                orig_name = orig_name[len(prefix):]
        dst = os.path.join(excel_dir, orig_name)
        shutil.copy2(src, dst)
        try: os.remove(src)
        except: pass
        if _excel_manager:
            _excel_manager._load_cache_from_files()
    return jsonify({"ok": True, "message": "已恢复"})


@app.route("/api/recycle/delete", methods=["POST"])
@require_admin
def api_recycle_delete():
    """永久删除回收站文件"""
    data = request.get_json() or {}
    filename = data.get("filename", "")
    mode = data.get("mode", "one")  # one or all
    if mode == "all":
        import glob
        for f in glob.glob(os.path.join(RECYCLE_DIR, "*")):
            try: os.remove(f)
            except: pass
        return jsonify({"ok": True, "message": "已清空回收站"})
    if not filename:
        return jsonify({"ok": False, "error": "请指定文件名"}), 400
    fp = os.path.join(RECYCLE_DIR, os.path.basename(filename))
    if os.path.exists(fp):
        os.remove(fp)
    return jsonify({"ok": True, "message": "已永久删除"})


# ---------- 密码验证 ----------

@app.route("/api/verify_password", methods=["POST"])
def api_verify_password():
    """验证管理员密码"""
    data = request.get_json() or {}
    pwd = data.get("password", "")
    users = load_users()
    admin = users.get("admin", {})
    if admin.get("password") == pwd:
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "密码错误"}), 403




@app.route("/api/restart", methods=["POST"])
@require_admin
def api_restart():
    """重启服务：延时3秒后启动新进程"""
    import subprocess as _sp, tempfile, sys as _sys
    cwd = os.getcwd()
    script = os.path.join(tempfile.gettempdir(), '_restart_shuaka.py')
    with open(script, 'w') as f:
        f.write(r'''
import time, os, sys, subprocess
time.sleep(3)
os.chdir({cwd!r})
kw = {{'stdout': subprocess.DEVNULL, 'stderr': subprocess.DEVNULL}}
if os.name == 'nt':
    kw['creationflags'] = 0x00000008
else:
    kw['start_new_session'] = True
subprocess.Popen({cmd!r}, **kw)
'''.format(cwd=cwd, cmd=[_sys.executable] + _sys.argv))
    # 启动重启脚本
    kw = dict(stdout=_sp.DEVNULL, stderr=_sp.DEVNULL)
    if os.name == 'nt':
        kw['creationflags'] = 0x00000008
    else:
        kw['start_new_session'] = True
    _sp.Popen([_sys.executable, script], **kw)
    # 延迟退出，让响应发出去
    def _exit():
        time.sleep(1)
        os._exit(0)
    threading.Thread(target=_exit, daemon=True).start()
    return jsonify({"ok": True, "message": "服务重启中，3秒后刷新页面"})


@app.route("/api/validate-data-path")
def api_validate_data_path():
    """验证数据路径是否有效，返回状态和建议"""
    path = request.args.get("path", "").strip()
    if not path:
        return jsonify({"ok": True, "status": "default", "message": "将使用默认路径"})
    if not os.path.exists(path):
        parent = os.path.dirname(path)
        can_create = os.path.isdir(parent)
        return jsonify({"ok": True, "status": "missing",
                        "message": "目录不存在", "can_create": can_create})
    if not os.path.isdir(path):
        return jsonify({"ok": True, "status": "invalid", "message": "路径不是有效目录"})
    files = _glob.glob(os.path.join(path, "签到记录_*.xlsx"))
    if not files:
        return jsonify({"ok": True, "status": "empty",
                        "message": "该目录下暂无签到数据"})
    return jsonify({"ok": True, "status": "ok",
                    "message": f"找到 {len(files)} 个签到数据文件", "files": len(files)})


@app.route("/api/create-data-dirs", methods=["POST"])
@require_admin
def api_create_data_dirs():
    """在指定路径创建数据目录结构和初始文件"""
    data = request.get_json() or {}
    path = data.get("path", "").strip()
    if not path:
        return jsonify({"ok": False, "error": "请指定路径"}), 400
    try:
        os.makedirs(path, exist_ok=True)
        # 创建初始签到 Excel（空模板）
        excel_path = os.path.join(path, "签到记录_1楼大厅.xlsx")
        if not os.path.exists(excel_path):
            try:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active; ws.title = "签到记录"
                ws.append(["序号", "姓名", "身份证号", "签到时间", "签到地点", "状态"])
                wb.save(excel_path)
            except Exception:
                pass  # Excel creation is optional
        return jsonify({"ok": True, "message": f"目录已创建: {path}"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/shutdown", methods=["POST"])
def api_shutdown():
    """安全关闭整个系统"""
    import sys as _sys
    def _do_shutdown():
        time.sleep(0.5)
        _sys.stdout.flush()
        _sys.stderr.flush()
        os._exit(0)
    threading.Thread(target=_do_shutdown, daemon=True).start()
    return jsonify({"ok": True, "message": "系统正在关闭"})

@app.route("/api/code-bundle")
def api_code_bundle():
    """返回所有可同步代码文件的 zip 包（供远程客户端拉取升级）"""
    import zipfile, io, glob as _g

    buf = io.BytesIO()
    # 远程升级只同步代码文件（DLL 驱动 20+MB 不随代码变动，走局域网同步即可）
    sync_ext = ('.py', '.ini', '.bat', '.yaml', '.txt', '.md', '.ico', '.png',
                '.html', '.css', '.js')
    sync_files = ('config.yaml', 'VERSION')
    exclude_files = {'settings.json', 'users.json', 'machine.json', '.syncing', '.gitignore'}

    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # 根目录文件
        for fn in os.listdir(BASE_DIR):
            fp = os.path.join(BASE_DIR, fn)
            if not os.path.isfile(fp):
                continue
            # 跳过数据/缓存文件
            if fn in exclude_files or fn.startswith('~$') or fn.startswith('.'):
                continue
            ext = os.path.splitext(fn)[1].lower()
            if ext in sync_ext or fn in sync_files:
                zf.write(fp, fn)

        # tablet/ 目录（全部文件，与局域网同步一致）
        for fn in os.listdir(TABLET_DIR):
            fp = os.path.join(TABLET_DIR, fn)
            if os.path.isfile(fp):
                zf.write(fp, f"tablet/{fn}")

    buf.seek(0)
    from flask import send_file
    return send_file(buf, mimetype='application/zip',
                     as_attachment=True, download_name='shuaka_code.zip')


@app.route("/api/sync_lan", methods=["POST"])
def api_sync_lan():
    """局域网同步：从 Mac SMB 共享拉取最新代码"""
    try:
        import platform as _plat, shutil as _sh, glob as _g
        if _plat.system() != 'Windows':
            return jsonify({"ok": False, "error": "仅支持 Windows 客户端"})

        local_dir = os.path.join(os.environ.get('LOCALAPPDATA', os.path.expanduser('~')), 'shuaka')
        src = r'\\192.168.50.226\shuaka'

        if not os.path.exists(os.path.join(src, 'main.py')):
            return jsonify({"ok": False, "error": f"无法访问 Mac SMB: {src}"})

        count = 0; errs = []; os.makedirs(local_dir, exist_ok=True)
        SYNC_EXT = ('.py', '.dll', '.ini', '.bat', '.yaml', '.txt', '.md', '.ico', '.png')
        SYNC_FILES = ('config.yaml', 'VERSION')
        for fn in os.listdir(src):
            ext = os.path.splitext(fn)[1].lower()
            if ext in SYNC_EXT or fn in SYNC_FILES:
                sp, dp = os.path.join(src, fn), os.path.join(local_dir, fn)
                if os.path.isfile(sp):
                    try: _sh.copy2(sp, dp); count += 1
                    except: errs.append(fn)
        ts, td = os.path.join(src, 'tablet'), os.path.join(local_dir, 'tablet')
        if os.path.isdir(ts):
            os.makedirs(td, exist_ok=True)
            for fn in os.listdir(ts):
                sf, df = os.path.join(ts, fn), os.path.join(td, fn)
                if os.path.isfile(sf):
                    try: _sh.copy2(sf, df); count += 1
                    except: errs.append(f"tablet/{fn}")

        total = count + len(errs)
        # 清理 Python 缓存
        for cache_dir in _g.glob(os.path.join(local_dir, '__pycache__')):
            try: _sh.rmtree(cache_dir)
            except: pass
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        _monitor_state['last_sync'] = now_str
        status = "✓ 完整" if not errs else f"⚠ 缺 {len(errs)} 个"
        _monitor_state['sync_result'] = f'{count}/{total} {status}'
        _write_event("sync_lan", text=f"局域网同步 {count}/{total} {status}")

        msg = f"局域网同步 {count}/{total} 文件 {status}"
        if errs: msg += f"，失败: {', '.join(errs[:3])}"
        return jsonify({"ok": True, "message": msg})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/sync_remote", methods=["POST"])
def api_sync_remote():
    """远程升级：从 Mac 服务器拉取最新代码（HTTP 方式，无需 Git）"""
    import urllib.request, zipfile, io, shutil, glob as _g, platform as _plat

    data = request.get_json(silent=True) or {}
    server_url = data.get("server_url", "").strip().rstrip('/')

    # 确定服务器地址：参数 > settings > 默认 LAN
    if not server_url:
        server_url = load_settings().get("sync_server_url", "").strip()
    if not server_url:
        server_url = os.environ.get("SYNC_SERVER_URL", "").strip()
    if not server_url:
        server_url = "http://192.168.50.226:5002"

    # 确定本地安装目录
    if _plat.system() == 'Windows':
        local_dir = os.path.join(os.environ.get('LOCALAPPDATA', os.path.expanduser('~')), 'shuaka')
    else:
        local_dir = os.path.dirname(os.path.abspath(__file__))

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def _download_with_powershell(url):
        """PowerShell 下载（Windows 最可靠，SSL 由系统处理）"""
        import subprocess as _sp, tempfile
        tmp = tempfile.mktemp(suffix='.zip')
        try:
            ps_cmd = (
                f'[Net.ServicePointManager]::SecurityProtocol = [Net.SecurityProtocolType]::Tls12;'
                f'$req = Invoke-WebRequest -Uri \"{url}\" -Headers @{{'
                f'\"ngrok-skip-browser-warning\"=\"true\";'
                f'\"User-Agent\"=\"ShuakaSync/2.0\"'
                f'}} -TimeoutSec 120 -UseBasicParsing;'
                f'[IO.File]::WriteAllBytes(\"{tmp}\", $req.Content)'
            )
            _sp.run(['powershell', '-WindowStyle', 'Hidden', '-NonInteractive', '-Command', ps_cmd],
                    check=True, timeout=130, capture_output=True)
            with open(tmp, 'rb') as f:
                return f.read()
        finally:
            try: os.remove(tmp)
            except: pass

    def _download_with_curl(url):
        """curl 下载"""
        import subprocess as _sp, tempfile
        tmp = tempfile.mktemp(suffix='.zip')
        try:
            _sp.run(['curl', '-s', '-o', tmp, '-H', 'ngrok-skip-browser-warning: true',
                     '-H', 'User-Agent: ShuakaSync/2.0',
                     '--connect-timeout', '10', '--max-time', '120', '-k', url],
                    check=True, timeout=130, capture_output=True)
            with open(tmp, 'rb') as f:
                return f.read()
        finally:
            try: os.remove(tmp)
            except: pass

    bundle_url = f"{server_url}/api/code-bundle"
    zip_data = None
    last_error = ""

    # Windows 上 PowerShell 处理 SSL 最可靠，优先使用
    is_win = _plat.system() == 'Windows'
    methods = [('powershell', _download_with_powershell), ('curl', _download_with_curl)] if is_win \
         else [('curl', _download_with_curl)]
    for name, fn in methods:
        try:
            zip_data = fn(bundle_url)
            break
        except Exception as e:
            last_error = f"{name}: {e}"
            continue

    if zip_data is None:
        return jsonify({"ok": False, "error": f"下载失败: {last_error}"})

    try:
        # 解压到本地
        os.makedirs(local_dir, exist_ok=True)
        count = 0
        with zipfile.ZipFile(io.BytesIO(zip_data)) as zf:
            for member in zf.namelist():
                if member.endswith('/'):
                    continue
                target = os.path.join(local_dir, member)
                os.makedirs(os.path.dirname(target), exist_ok=True)
                with zf.open(member) as src, open(target, 'wb') as dst:
                    shutil.copyfileobj(src, dst)
                count += 1

        # 清理 Python 缓存
        for cache_dir in _g.glob(os.path.join(local_dir, '__pycache__')):
            try: shutil.rmtree(cache_dir)
            except: pass
        tablet_cache = os.path.join(local_dir, 'tablet', '__pycache__')
        if os.path.isdir(tablet_cache):
            try: shutil.rmtree(tablet_cache)
            except: pass

        _monitor_state['last_sync'] = now_str
        _monitor_state['sync_result'] = f'远程: {count} 文件 来自 {server_url}'
        msg = f"远程升级完成: {count} 个文件已更新 (来源: {server_url})"
        _write_event("sync_lan", text=msg)
        return jsonify({"ok": True, "message": msg, "files": count, "server": server_url})

    except Exception as e:
        return jsonify({"ok": False, "error": f"解压失败: {str(e)[:120]}"})


@app.route("/api/sync_trigger", methods=["POST"])
def api_sync_trigger():
    """桌面控制面板同步按钮入口 → 转发到 sync_remote"""
    return api_sync_remote()


@app.route("/desktop")
def serve_desktop():
    """桌面控制面板页面"""
    return send_from_directory('tablet', 'desktop.html')

@app.route("/api/browse-dirs")
def api_browse_dirs():
    """浏览本地目录，用于文件夹选择器"""
    path = request.args.get("path", "").strip()
    if not path:
        if os.name == 'nt':
            roots = [{"name": d + ":\\", "path": d + ":\\", "type": "drive"}
                     for d in 'CDEFGH' if os.path.exists(d + ":\\")]
            return jsonify({"ok": True, "path": "", "items": roots})
        else:
            path = os.path.expanduser("~")
    if not os.path.isdir(path):
        return jsonify({"ok": False, "error": "路径不存在"}), 400
    items = []
    try:
        for name in sorted(os.listdir(path)):
            full = os.path.join(path, name)
            if os.path.isdir(full) and not name.startswith('.'):
                items.append({"name": name, "path": full, "type": "dir"})
    except PermissionError:
        return jsonify({"ok": False, "error": "无权限访问"}), 403
    parent = os.path.dirname(path)
    if parent != path:
        items.insert(0, {"name": "..", "path": parent, "type": "parent"})
    return jsonify({"ok": True, "path": path, "items": items})


@app.route("/api/change_password", methods=["POST"])
def api_change_password():
    """修改用户密码"""
    data = request.get_json() or {}
    username = data.get("username", "").strip()
    new_password = data.get("new_password", "").strip()
    if not username or not new_password:
        return jsonify({"ok": False, "error": "用户名和密码不能为空"}), 400
    if len(new_password) < 6:
        return jsonify({"ok": False, "error": "密码至少6位"}), 400
    users = load_users()
    if username not in users:
        return jsonify({"ok": False, "error": "用户不存在"}), 404
    users[username]["password"] = new_password
    save_users(users)
    return jsonify({"ok": True, "message": f"用户 {username} 密码已修改"})

def start_server(host="0.0.0.0", port=5002):
    _load_sessions()  # 恢复上次的登录会话

    local_ip = get_local_ip()
    all_ips = get_all_ips()

    detect_ngrok_url()  # 自动检测已运行的 ngrok 隧道

    print(f"\n{'='*55}")
    _start_event_polling()
    print(f"  叫号大屏 Web 服务器已启动")
    print(f"  管理后台: http://{local_ip}:{port}/admin")
    print(f"  默认账号: admin / admin123")
    print(f"  大屏展示: http://{local_ip}:{port}/")
    print(f"  {'─'*45}")
    print(f"  内网访问地址:")
    for ip in all_ips:
        print(f"    http://{ip}:{port}")
    if _ngrok_url:
        print(f"  {'─'*45}")
        print(f"  外网访问地址: {_ngrok_url}")
    print(f"{'='*55}\n")

    t = threading.Thread(
        target=lambda: app.run(host=host, port=port, debug=False, use_reloader=False),
        daemon=True
    )
    t.start()
    return t
