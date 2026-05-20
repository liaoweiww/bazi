"""
Excel 签到记录管理器
- 按地点分文件存储，避免百度网盘同步冲突
- 线程安全写入
- 支持读取所有地点文件合并展示
"""

import os
import re
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook

class ExcelManager:
    def __init__(self, excel_dir, location):
        self.excel_dir = os.path.abspath(excel_dir)
        self.location = location
        self.lock = threading.Lock()
        self._records_cache = []  # 所有记录的内存缓存

        os.makedirs(self.excel_dir, exist_ok=True)

    def _filepath(self, loc=None):
        """获取指定地点的 Excel 文件路径"""
        loc = loc or self.location
        safe_loc = re.sub(r'[\\/:*?"<>|]', '_', loc)
        return os.path.join(self.excel_dir, f"签到记录_{safe_loc}.xlsx")

    def _ensure_file(self, filepath):
        """确保 Excel 文件存在且有表头"""
        if not os.path.exists(filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = "签到记录"
            ws.append(["序号", "姓名", "身份证号", "签到时间", "签到地点", "状态"])
            # 设置列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 22
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 10
            wb.save(filepath)

    def add_record(self, name, id_number, sign_time=None, status="等待中", extra=None, _rebuild=False):
        """
        添加一条签到记录。_rebuild=True 时只写Excel不更新缓存。
        返回: 记录字典
        """
        sign_time = sign_time or datetime.now()
        time_str = sign_time.strftime("%Y-%m-%d %H:%M:%S")

        with self.lock:
            filepath = self._filepath()
            self._ensure_file(filepath)

            wb = load_workbook(filepath)
            ws = wb.active

            seq = ws.max_row

            record = {
                "seq": seq,
                "name": name,
                "id_number": id_number,
                "sign_time": time_str,
                "location": self.location,
                "status": status
            }

            ws.append([seq, name, id_number, time_str, self.location, status])
            wb.save(filepath)

            if extra:
                record.update(extra)

            if not _rebuild:
                self._records_cache.append(record)
                self._records_cache.sort(key=lambda r: r["sign_time"], reverse=True)

            return record

    def get_all_records(self):
        """
        读取所有地点的签到记录（合并）
        包括本机记录和百度网盘同步过来的异地记录
        """
        records = []
        try:
            for filename in sorted(os.listdir(self.excel_dir)):
                if not filename.endswith('.xlsx'):
                    continue
                # 跳过临时文件
                if filename.startswith('~$'):
                    continue

                filepath = os.path.join(self.excel_dir, filename)
                try:
                    wb = load_workbook(filepath, read_only=True, data_only=True)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0] is None:
                            continue
                        records.append({
                            "seq": row[0],
                            "name": str(row[1]) if row[1] else "",
                            "id_number": str(row[2]) if row[2] else "",
                            "sign_time": str(row[3]) if row[3] else "",
                            "location": str(row[4]) if row[4] else "",
                            "status": str(row[5]) if row[5] else "等待中"
                        })
                    wb.close()
                except Exception:
                    continue
        except Exception:
            pass

        # 按签到时间倒序排列
        records.sort(key=lambda r: r["sign_time"], reverse=True)
        return records

    def get_today_records(self):
        """获取今日签到记录"""
        today = datetime.now().strftime("%Y-%m-%d")
        return [r for r in self.get_all_records()
                if r["sign_time"].startswith(today)]

    def update_status(self, seq, location, new_status):
        """更新某条记录的状态"""
        with self.lock:
            filepath = self._filepath(location)
            if not os.path.exists(filepath):
                return False

            wb = load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                if row[0].value == seq:
                    row[5].value = new_status
                    wb.save(filepath)

                    # 更新缓存
                    for r in self._records_cache:
                        if r["seq"] == seq and r["location"] == location:
                            r["status"] = new_status
                            break
                    return True
            wb.close()
            return False
